import pandas as pd
import os
from openpyxl import load_workbook

log_file = f"{os.getcwd()}/src/testing/Agent_Log.xlsx"

def prepare_entry(data=[], columns=[], log_file="app.log"):
    log_data = pd.DataFrame(data)
    log_data = log_data.transpose()
    log_data.columns=columns

    log_data["Timestamp"] = pd.to_datetime("now")
    return log_data

def add_to_log(new_row: pd.DataFrame, log_data: pd.DataFrame = pd.DataFrame()):
    log_data = pd.concat([log_data, pd.DataFrame([new_row])], ignore_index=True)
    return log_data

def update_log(log_data: pd.DataFrame, row_id, column_name, new_value):
    log_data.at[row_id, column_name] = new_value
    return log_data

def append_df_to_excel(log_df: pd.DataFrame, filename: str = log_file, sheet_name='Sheet1'):
    """
    Appends a DataFrame [df] to an existing Excel file [filename].
    If the file does not exist, it will be created.
    """
    
    if not os.path.isfile(filename):
        # File does not exist, create it with df
        log_df.to_excel(filename, sheet_name=sheet_name, index=False)
    else:
        # File exists, append without duplicating headers
        wb = load_workbook(filename)
        if sheet_name in wb.sheetnames:
            startrow = wb[sheet_name].max_row
        else:
            startrow = 0
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            log_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)